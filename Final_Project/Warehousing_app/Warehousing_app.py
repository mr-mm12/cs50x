# Some parts were completed using artificial 
# intelligence, and in front of that part it says "AI".
import csv
import json
import logging
import os
import threading
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ------------------------- Logging Configuration -------------------------
# Setup basic logging to track events and debug issues
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("warehouse")

# ------------------------- Constants -------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))  # Directory where the script resides
SOUND_FILENAME = "clicked.wav"  # Default click sound file
SOUND_PATH = os.path.join(SCRIPT_DIR, SOUND_FILENAME)

_sound_play_fn = None 

# ------------------------- Sound Playback Setup -------------------------
# Attempt to import platform-specific or third-party sound playback libraries

# ------------------------- Sound Playback Setup -------------------------
# Attempt to import platform-specific or third-party sound playback libraries.
# This ensures that a click sound can be played asynchronously, depending on
# which library is available on the system.

try:
    import winsound  # Windows built-in library for playing sounds

    def _winsound_play(path: str):
        """
        Play a WAV file using the built-in Windows winsound library asynchronously.
        
        Args:
            path (str): Full path to the WAV file to be played.
        """
        winsound.PlaySound(path, winsound.SND_FILENAME | winsound.SND_ASYNC)

    _sound_play_fn = _winsound_play  # Assign the function pointer
    logger.debug("Using winsound for click sounds")
except Exception:
    logger.debug("winsound not available")

# If winsound is not available, try using simpleaudio (cross-platform)
if _sound_play_fn is None:
    try:
        import simpleaudio as sa

        def _simpleaudio_play(path: str):
            """
            Play a WAV file using the simpleaudio library.

            Args:
                path (str): Full path to the WAV file to be played.
            """
            wave_obj = sa.WaveObject.from_wave_file(path)
            wave_obj.play()  # This starts playback asynchronously

        _sound_play_fn = _simpleaudio_play
        logger.debug("Using simpleaudio for click sounds")
    except Exception:
        logger.debug("simpleaudio not available")

# If neither winsound nor simpleaudio is available, fallback to playsound
if _sound_play_fn is None:
    try:
        from playsound import playsound

        def _playsound_play(path: str):
            """
            Play a WAV file using the playsound library (blocking call).

            Args:
                path (str): Full path to the WAV file to be played.
            """
            playsound(path)

        def _playsound_async(path: str):
            """
            Play a WAV file asynchronously in a separate thread using playsound.

            Args:
                path (str): Full path to the WAV file to be played.
            """
            threading.Thread(target=_playsound_play, args=(path,), daemon=True).start()

        _sound_play_fn = _playsound_async
        logger.debug("Using playsound for click sounds")
    except Exception:
        logger.debug("playsound not available")



def play_click_sound_async() -> None:  # "AI"
    """
    Play a click sound asynchronously if available.
    
    This function attempts to play a click sound in a non-blocking way.
    It checks if the sound file exists and if a playback function is available.
    Any errors during playback are logged for debugging but do not raise exceptions,
    ensuring that sound failures do not interrupt program execution.
    """
    # If the sound file does not exist or no sound function is available, do nothing
    if not os.path.isfile(SOUND_PATH) or _sound_play_fn is None:
        return

    def worker():
        """
        Worker thread to handle the asynchronous playback.
        Encapsulates the call to the platform-specific sound function
        and logs any exceptions without crashing the program.
        """
        try:
            _sound_play_fn(SOUND_PATH)  # Call the assigned sound playback function
        except Exception as exc:
            logger.debug("Failed to play sound: %s", exc)

    # Start the playback in a daemon thread so it runs asynchronously
    threading.Thread(target=worker, daemon=True).start()



# ------------------------- Product Class -------------------------
class Product:
    """Represents a single product with quantity and transaction history.

    This class tracks the stock of a product and maintains a record
    of all transactions (additions, sales, replacements) for auditing.
    """

    def __init__(self, name: str, quantity: int):
        """
        Initialize a Product instance.
        
        name: Name of the product (normalized to lowercase for consistency)
        quantity: Initial stock quantity
        """
        self.name: str = name.lower()  # Normalize name to lowercase
        self.quantity: int = int(quantity)
        self.transactions: List[Dict] = []  # Stores all actions with timestamps
        self._record("initial", self.quantity)  # Record initial stock

    def _record(self, action: str, qty: int) -> None:
        """
        Record a transaction in the product's history.
        
        action: Type of operation ('add', 'sell', 'initial', etc.)
        qty: Quantity affected by this action
        """
        self.transactions.append(
            {
                "action": action,
                "quantity": int(qty),
                "datetime": datetime.now().isoformat(timespec="seconds"),
            }
        )

    def add(self, qty: int) -> None:
        """
        Increase the product's quantity and record the transaction.
        
        qty: Number of units added to stock
        """
        qty = int(qty)
        self.quantity += qty
        self._record("add", qty)

    def sell(self, qty: int) -> bool:
        """
        Reduce the product's quantity if enough stock exists.
        
        qty: Number of units to sell
        Returns True if sale succeeded, False otherwise.
        """
        qty = int(qty)
        if qty <= self.quantity:
            self.quantity -= qty
            self._record("sell", qty)
            return True
        return False

    def replace_initial(self, qty: int) -> None:
        """
        Replace the current stock with a new quantity and record it.
        
        qty: New stock quantity
        """
        qty = int(qty)
        self.quantity = qty
        self._record("initial_replace", qty)


# ------------------------- Inventory Class -------------------------
class Inventory:
    """
    Inventory manager for multiple products.

    Handles adding, selling, removing products, undo/redo operations,
    CSV import/export, and autosaving to disk.
    """

    def __init__(self, autosave_filename: Optional[str] = None):
        """
        Initialize Inventory.

        autosave_filename: Optional path for autosave JSON file.
        """
        self.products: Dict[str, Product] = {}  # Dictionary of products by name
        self.undo_stack: List[Dict] = []        # Stack for undo operations
        self.redo_stack: List[Dict] = []        # Stack for redo operations
        self.autosave_file = (
            autosave_filename
            if autosave_filename
            else os.path.join(SCRIPT_DIR, "warehouse_data.json")  # Default file path
        )
        try:
            self.load()  # Load saved inventory if available
        except Exception as exc:
            logger.info("No autosave loaded: %s", exc)

    # ----------------- Product Management -----------------
    def list_products(self) -> List[Product]:
        """Return a list of all products currently in inventory."""
        return list(self.products.values())

    def add_or_replace_product(self, name: str, qty: int) -> Tuple[bool, Optional[int]]:
        """
        Add a new product or replace existing product quantity.

        Returns:
            (True, None) if new product added,
            (False, old_qty) if existing product quantity replaced.
        """
        name = name.lower()
        qty = int(qty)
        if name in self.products:
            old = self.products[name].quantity
            self.products[name].replace_initial(qty)
            self.undo_stack.append({"op": "replace", "name": name, "old": old, "new": qty})
            self.redo_stack.clear()
            self.save()
            return False, old
        else:
            p = Product(name, qty)
            self.products[name] = p
            self.undo_stack.append({"op": "add_product", "name": name, "qty": qty})
            self.redo_stack.clear()
            self.save()
            return True, None

    def add_stock(self, name: str, qty: int) -> bool:
        """
        Add stock to an existing product.

        Returns True if successful, False if product does not exist.
        """
        name = name.lower()
        qty = int(qty)
        if name not in self.products:
            return False
        prod = self.products[name]
        prev = prod.quantity
        prod.add(qty)
        self.undo_stack.append({"op": "add", "name": name, "qty": qty, "prev": prev})
        self.redo_stack.clear()
        self.save()
        return True

    def sell_stock(self, name: str, qty: int) -> bool:
        """
        Sell a quantity of an existing product.

        Returns True if sale succeeded, False otherwise.
        """
        name = name.lower()
        qty = int(qty)
        if name not in self.products:
            return False
        prod = self.products[name]
        prev = prod.quantity
        ok = prod.sell(qty)
        if not ok:
            return False
        self.undo_stack.append({"op": "sell", "name": name, "qty": qty, "prev": prev})
        self.redo_stack.clear()
        self.save()
        return True

    def remove_product(self, name: str) -> bool:
        """
        Remove a product completely from inventory.

        Returns True if product existed and was removed, False otherwise.
        """
        name = name.lower()
        if name in self.products:
            prod = self.products.pop(name)
            self.undo_stack.append({"op": "remove_product", "product": self._serialize_product(prod)})
            self.redo_stack.clear()
            self.save()
            return True
        return False

    # ----------------- Undo/Redo -----------------
    def undo(self) -> bool:  # "AI"
        """Undo the last operation if possible."""
        if not self.undo_stack:
            return False
        op = self.undo_stack.pop()
        try:
            self._apply_undo(op)
            self.redo_stack.append(op)
            self.save()
            return True
        except Exception as exc:
            logger.exception("Undo failed: %s", exc)
            return False

    def redo(self) -> bool:  # "AI"
        """Redo the last undone operation if possible."""
        if not self.redo_stack:
            return False
        op = self.redo_stack.pop()
        try:
            self._apply_redo(op)
            self.undo_stack.append(op)
            self.save()
            return True
        except Exception as exc:
            logger.exception("Redo failed: %s", exc)
            return False

    # ----------------- Internal Undo/Redo Helpers -----------------
    def _apply_undo(self, op: Dict) -> None:
        """Apply a single undo operation based on stored data."""
        typ = op.get("op")
        if typ == "add_product":
            name = op["name"]
            if name in self.products:
                self.products.pop(name)
        elif typ == "remove_product":
            prod_data = op["product"]
            p = self._deserialize_product(prod_data)
            self.products[p.name] = p
        elif typ == "replace":
            name = op["name"]
            old = op["old"]
            if name in self.products:
                self.products[name].replace_initial(old)
        elif typ == "add":
            name = op["name"]
            prev = op["prev"]
            if name in self.products:
                self.products[name].quantity = prev
                self.products[name]._record("undo_add", op.get("qty", 0))
        elif typ == "sell":
            name = op["name"]
            prev = op["prev"]
            if name in self.products:
                self.products[name].quantity = prev
                self.products[name]._record("undo_sell", op.get("qty", 0))
        else:
            logger.debug("Unknown undo op: %s", op)

    def _apply_redo(self, op: Dict) -> None:
        """Apply a single redo operation based on stored data."""
        typ = op.get("op")
        if typ == "add_product":
            name = op["name"]
            qty = op.get("qty", 0)
            if name not in self.products:
                self.products[name] = Product(name, qty)
        elif typ == "remove_product":
            prod_data = op["product"]
            name = prod_data["name"]
            if name in self.products:
                self.products.pop(name)
        elif typ == "replace":
            name = op["name"]
            new = op.get("new")
            if name in self.products and new is not None:
                self.products[name].replace_initial(new)
        elif typ == "add":
            name = op["name"]
            qty = op.get("qty", 0)
            if name in self.products:
                self.products[name].add(qty)
        elif typ == "sell":
            name = op["name"]
            qty = op.get("qty", 0)
            if name in self.products:
                self.products[name].sell(qty)
        else:
            logger.debug("Unknown redo op: %s", op)

    # ----------------- Serialization Helpers -----------------
    def _serialize_product(self, p: Product) -> Dict:
        """Convert a Product instance to a dictionary for saving or undo/redo."""
        return {"name": p.name, "quantity": p.quantity, "transactions": p.transactions.copy()}

    def _deserialize_product(self, data: Dict) -> Product:
        """Convert a dictionary back into a Product instance."""
        p = Product(data["name"], 0)
        p.quantity = int(data.get("quantity", 0))
        p.transactions = data.get("transactions", []).copy()
        return p

    # ----------------- Save / Load -----------------
    def save(self) -> bool:  # "AI"
        """Save inventory data to JSON file on disk."""
        try:
            data = {"products": [self._serialize_product(p) for p in self.products.values()]}
            with open(self.autosave_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as exc:
            logger.exception("Failed to save inventory: %s", exc)
            return False

    def load(self) -> bool:
        """Load inventory data from JSON file on disk."""
        try:
            if not os.path.isfile(self.autosave_file):
                return False
            with open(self.autosave_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            prods = data.get("products", [])
            self.products.clear()
            for pd in prods:
                p = self._deserialize_product(pd)
                self.products[p.name] = p
            return True
        except Exception as exc:
            logger.exception("Failed to load inventory: %s", exc)
            self.products = {}
            return False

    # ----------------- CSV Import -----------------
    def import_from_csv(self, path: str) -> Tuple[int, int]:  # "AI"
        """
        Import products from a CSV file.

        Returns:
            (added_count, replaced_count) indicating number of new products added
            and existing products replaced.
        """
        pairs = self.parse_csv_file(path)
        added = replaced = 0
        for name, qty in pairs:
            name = name.lower()
            if name in self.products:
                old = self.products[name].quantity
                self.products[name].replace_initial(qty)
                self.undo_stack.append({"op": "replace", "name": name, "old": old, "new": qty})
                replaced += 1
            else:
                p = Product(name, qty)
                self.products[name] = p
                self.undo_stack.append({"op": "add_product", "name": name, "qty": qty})
                added += 1
        self.redo_stack.clear()
        self.save()
        return added, replaced

    @staticmethod
    def parse_csv_file(path: str) -> List[Tuple[str, int]]:  # "AI"
        """
        Parse CSV file to extract product name and quantity pairs.

        Returns a list of tuples: (product_name, quantity)
        """
        pairs: List[Tuple[str, int]] = []
        try:
            with open(path, newline="", encoding="utf-8") as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    if not row:
                        continue
                    cells = [c.strip() for c in row if c and str(c).strip() != ""]
                    if not cells:
                        continue
                    low0 = cells[0].lower()
                    if low0 in ("product", "name") and len(cells) > 1 and not Inventory._is_int(cells[1]):
                        continue
                    qty = None
                    name = None
                    for i in range(1, len(cells)):
                        if Inventory._is_int(cells[i]):
                            qty = int(float(cells[i]))
                            name = cells[0]
                            break
                    if qty is None and len(cells) == 1:
                        parts = cells[0].replace(",", " ").split()
                        if len(parts) >= 2 and Inventory._is_int(parts[-1]):
                            qty = int(float(parts[-1]))
                            name = " ".join(parts[:-1])
                    if qty is None:
                        for i, c in enumerate(cells):
                            if Inventory._is_int(c):
                                qty = int(float(c))
                                name = cells[i - 1] if i > 0 else cells[0]
                                break
                    if qty is not None and name:
                        pairs.append((name.strip(), int(qty)))
        except Exception as exc:
            logger.exception("Failed to parse CSV %s: %s", path, exc)
        return pairs

    @staticmethod
    def _is_int(s: str) -> bool:
        """Check if a string can be converted to an integer."""
        try:
            _ = int(float(s))
            return True
        except Exception:
            return False


# ------------------------- WarehouseApp GUI Class -------------------------
class WarehouseApp(tk.Tk):  # "AI"
    """Tkinter GUI to interact with Inventory class for managing products and transactions."""

    def __init__(self, inventory: Inventory):
        """Initialize main window, attach Inventory, and build GUI components."""
        super().__init__()
        self.inventory = inventory
        self.title("Warehouse Manager — Improved")
        self.geometry("1000x640")
        self.minsize(800, 420)

        # Build GUI sections
        self._build_menu()         # top menu (File/Edit)
        self._build_top()          # product input and search
        self._build_middle()       # product list and transactions
        self._build_bottom()       # bottom buttons
        self._bind_shortcuts()     # keyboard shortcuts
        self.refresh_products_table()  # populate table with inventory

    def _build_menu(self):
        """Create menu bar with File and Edit options."""
        menubar = tk.Menu(self)
        filem = tk.Menu(menubar, tearoff=False)
        filem.add_command(label="Import CSV...", command=self._wrap(self.on_import_csv))
        filem.add_command(label="Export to Excel...", command=self._wrap(self.on_export_excel))
        filem.add_separator()
        filem.add_command(label="Save now", command=self._wrap(self._save_now))
        filem.add_command(label="Load from disk", command=self._wrap(self._manual_load))
        filem.add_separator()
        filem.add_command(label="Quit", command=self._wrap(self.quit))
        menubar.add_cascade(label="File", menu=filem)

        editm = tk.Menu(menubar, tearoff=False)
        editm.add_command(label="Undo (Ctrl+Z)", command=self._wrap(self.on_undo))
        editm.add_command(label="Redo (Ctrl+Y)", command=self._wrap(self.on_redo))
        menubar.add_cascade(label="Edit", menu=editm)

        self.config(menu=menubar)

    def _build_top(self):  # "AI"
        """Create the top frame for product input, quantity, buttons, and search bar."""
        top = ttk.Frame(self, padding=8)
        top.pack(side=tk.TOP, fill=tk.X)

        ttk.Label(top, text="Product:").grid(row=0, column=0, sticky=tk.W)
        self.entry_name = ttk.Entry(top, width=30)
        self.entry_name.grid(row=0, column=1, padx=6)

        ttk.Label(top, text="Qty:").grid(row=0, column=2, sticky=tk.W)
        self.entry_qty = ttk.Entry(top, width=10)
        self.entry_qty.grid(row=0, column=3, padx=6)

        btn_add = ttk.Button(top, text="Add / Replace", command=self._wrap(self.on_add_replace_product))
        btn_add.grid(row=0, column=4, padx=6)

        btn_remove = ttk.Button(top, text="Remove", command=self._wrap(self.on_remove_product))
        btn_remove.grid(row=0, column=5, padx=6)

        ttk.Label(top, text="Search:").grid(row=0, column=6, sticky=tk.W, padx=(12, 0))
        self.entry_search = ttk.Entry(top, width=25)
        self.entry_search.grid(row=0, column=7, padx=6)
        self.entry_search.bind("<KeyRelease>", lambda e: self.refresh_products_table())

    def _build_middle(self):
        """Create middle section with product tree on left and transactions + controls on right."""
        middle = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        middle.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=8, pady=6)

        # Left frame: product list
        left_frame = ttk.Frame(middle)
        middle.add(left_frame, weight=1)
        ttk.Label(left_frame, text="Products").pack(anchor=tk.W)
        cols = ("name", "qty")
        self.tree = ttk.Treeview(left_frame, columns=cols, show="headings", selectmode="browse")
        self.tree.heading("name", text="Product")
        self.tree.heading("qty", text="Quantity")
        self.tree.column("name", width=360)
        self.tree.column("qty", width=100, anchor=tk.CENTER)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        vsb = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Right frame: selected product info, stock controls, transactions
        right_frame = ttk.Frame(middle, width=380)
        middle.add(right_frame, weight=0)
        ttk.Label(right_frame, text="Selected Product:").pack(anchor=tk.W)
        self.label_selected = ttk.Label(right_frame, text="(none)")
        self.label_selected.pack(anchor=tk.W, pady=(0, 8))

        # Stock controls
        ctrl = ttk.Frame(right_frame)
        ctrl.pack(anchor=tk.W, pady=4, fill=tk.X)
        ttk.Label(ctrl, text="Amount:").grid(row=0, column=0, padx=4)
        self.entry_amount = ttk.Entry(ctrl, width=12)
        self.entry_amount.grid(row=0, column=1, padx=4)
        btn_add_stock = ttk.Button(ctrl, text="Add", command=self._wrap(self.on_add_stock))
        btn_add_stock.grid(row=0, column=2, padx=6)
        btn_sell = ttk.Button(ctrl, text="Sell", command=self._wrap(self.on_sell_stock))
        btn_sell.grid(row=0, column=3, padx=6)

        btn_tx = ttk.Button(right_frame, text="Open Transactions Window", command=self._wrap(self.on_open_transactions))
        btn_tx.pack(anchor=tk.W, pady=(8, 0))

        # Transactions tree
        ttk.Label(right_frame, text="Transactions:").pack(anchor=tk.W, pady=(12, 0))
        tx_cols = ("action", "qty", "ts", "stock")
        self.tree_tx = ttk.Treeview(right_frame, columns=tx_cols, show="headings", height=12)
        for c, h in zip(tx_cols, ("Action", "Qty", "Date/Time", "Stock After")):
            self.tree_tx.heading(c, text=h)
        self.tree_tx.column("action", width=80, anchor=tk.CENTER)
        self.tree_tx.column("qty", width=60, anchor=tk.CENTER)
        self.tree_tx.column("ts", width=160)
        self.tree_tx.column("stock", width=80, anchor=tk.CENTER)
        self.tree_tx.pack(fill=tk.BOTH, expand=True)

    def _build_bottom(self):
        """Create bottom button panel for Export, Undo/Redo, Clear All, and Quit."""
        bottom = ttk.Frame(self, padding=6)
        bottom.pack(side=tk.BOTTOM, fill=tk.X)
        btn_export = ttk.Button(bottom, text="Export to Excel", command=self._wrap(self.on_export_excel))
        btn_export.pack(side=tk.LEFT, padx=6)
        btn_undo = ttk.Button(bottom, text="Undo", command=self._wrap(self.on_undo))
        btn_undo.pack(side=tk.LEFT, padx=6)
        btn_redo = ttk.Button(bottom, text="Redo", command=self._wrap(self.on_redo))
        btn_redo.pack(side=tk.LEFT, padx=6)
        btn_clear = ttk.Button(bottom, text="Clear All", command=self._wrap(self.on_clear_all))
        btn_clear.pack(side=tk.LEFT, padx=6)
        btn_quit = ttk.Button(bottom, text="Quit", command=self._wrap(self.quit))
        btn_quit.pack(side=tk.RIGHT, padx=6)

    def _wrap(self, fn):
        """Wrap a function to play a click sound asynchronously before execution."""
        def wrapped(*a, **kw):
            play_click_sound_async()
            return fn(*a, **kw)
        return wrapped

    def _bind_shortcuts(self):
        """Bind keyboard shortcuts for Undo (Ctrl+Z) and Redo (Ctrl+Y)."""
        self.bind_all("<Control-z>", lambda e: self._wrap(self.on_undo)())
        self.bind_all("<Control-y>", lambda e: self._wrap(self.on_redo)())

    def refresh_products_table(self):  # "AI"
        """Update the product treeview based on current inventory and search filter."""
        search = self.entry_search.get().strip().lower() if hasattr(self, "entry_search") else ""
        desired = [name for name in sorted(self.inventory.products.keys()) if (not search or search in name)]
        current = set(self.tree.get_children())
        desired_set = set(desired)

        # remove items no longer needed
        for iid in list(current - desired_set):
            self.tree.delete(iid)

        # add new items and update existing
        for name in desired:
            prod = self.inventory.products.get(name)
            if name in current:
                vals = self.tree.item(name, "values")
                try:
                    current_qty = int(vals[1])
                except Exception:
                    current_qty = None
                if current_qty != prod.quantity:
                    self.tree.item(name, values=(prod.name, prod.quantity))
            else:
                self.tree.insert("", tk.END, iid=name, values=(prod.name, prod.quantity))

    def _on_tree_select(self, event):
        """Handle product selection: display product info and transactions."""
        sel = self.tree.selection()
        if not sel:
            self._show_selected(None)
            return
        iid = sel[0]
        prod = self.inventory.products.get(iid)
        self._show_selected(prod)

    def _show_selected(self, prod: Optional[Product]) -> None:
        """Update the selected product label and populate its transaction history."""
        self.tree_tx.delete(*self.tree_tx.get_children())
        if prod is None:
            self.label_selected.config(text="(none)")
            self.selected_name = None
            return
        self.selected_name = prod.name
        self.label_selected.config(text=f"{prod.name} — {prod.quantity}")
        current = 0
        for entry in prod.transactions:
            act = entry.get("action")
            qty = entry.get("quantity")
            ts = entry.get("datetime")
            if act in ("initial", "initial_replace"):
                current = qty
            elif act == "add":
                current += qty
            elif act == "sell":
                current -= qty
            self.tree_tx.insert("", tk.END, values=(act, qty, ts, current))

    # ---------------- Product Operations ----------------
    def on_add_replace_product(self):  # "AI"
        """Add a new product or replace an existing one in inventory."""
        name = self.entry_name.get().strip()
        if not name:
            messagebox.showerror("Usage", "Please enter a product name.")
            return
        try:
            qty = int(self.entry_qty.get().strip())
            if qty < 0:
                raise ValueError("negative")
        except Exception:
            messagebox.showerror("Usage", "Quantity must be a non-negative integer.")
            return
        added, old = self.inventory.add_or_replace_product(name, qty)
        self.entry_name.delete(0, tk.END)
        self.entry_qty.delete(0, tk.END)
        self.refresh_products_table()
        msg = f"Added '{name}' with qty {qty}." if added else f"Replaced '{name}' old qty {old} -> {qty}."
        messagebox.showinfo("Product updated", msg)

    def on_remove_product(self):  # "AI"
        """Remove a product from inventory if it exists."""
        name = self.entry_name.get().strip()
        if not name:
            messagebox.showerror("Usage", "Please enter product name to remove.")
            return
        if self.inventory.remove_product(name):
            self.entry_name.delete(0, tk.END)
            self.refresh_products_table()
            messagebox.showinfo("Removed", f"Product '{name}' removed.")
        else:
            messagebox.showerror("Error", f"Product '{name}' not found.")

    def on_add_stock(self):  # "AI"
        """Increase stock of selected product by specified amount."""
        if not hasattr(self, "selected_name") or not self.selected_name:
            messagebox.showerror("Usage", "Select a product first.")
            return
        try:
            amt = int(self.entry_amount.get().strip())
            if amt <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Usage", "Enter a positive integer amount.")
            return
        ok = self.inventory.add_stock(self.selected_name, amt)
        if not ok:
            messagebox.showerror("Error", "Failed to add stock (product missing).")
            return
        self.entry_amount.delete(0, tk.END)
        self._show_selected(self.inventory.products[self.selected_name])
        self.refresh_products_table()

    def on_sell_stock(self):  # "AI"
        """Decrease stock of selected product by specified amount, ensuring sufficient quantity."""
        if not hasattr(self, "selected_name") or not self.selected_name:
            messagebox.showerror("Usage", "Select a product first.")
            return
        try:
            amt = int(self.entry_amount.get().strip())
            if amt <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Usage", "Enter a positive integer amount.")
            return
        ok = self.inventory.sell_stock(self.selected_name, amt)
        if not ok:
            messagebox.showerror("Error", "Not enough stock or product missing.")
            return
        self.entry_amount.delete(0, tk.END)
        self._show_selected(self.inventory.products[self.selected_name])
        self.refresh_products_table()

    # ---------------- Transactions ----------------
    def on_open_transactions(self):
        """Open a popup window displaying all transactions for the selected product."""
        if not hasattr(self, "selected_name") or not self.selected_name:
            messagebox.showerror("Usage", "Select a product first.")
            return
        prod = self.inventory.products[self.selected_name]
        popup = tk.Toplevel(self)
        popup.title(f"Transactions — {prod.name}")
        tv = ttk.Treeview(popup, columns=("action", "qty", "ts"), show="headings")
        tv.heading("action", text="Action")
        tv.heading("qty", text="Qty")
        tv.heading("ts", text="Date/Time")
        tv.pack(fill=tk.BOTH, expand=True)
        for e in prod.transactions:
            tv.insert("", tk.END, values=(e["action"], e["quantity"], e["datetime"]))

    def on_import_csv(self):  # "AI"
        """Load product data from a CSV file and update the table."""
        path = filedialog.askopenfilename(title="Open CSV", filetypes=[("CSV files", "*.csv"), ("All files", "*")])
        if not path:
            return
        try:
            added, replaced = self.inventory.import_from_csv(path)
            self.refresh_products_table()
            messagebox.showinfo("Import", f"Completed. Added: {added}, Replaced: {replaced}")
        except Exception as exc:
            logger.exception("Import CSV failed: %s", exc)
            messagebox.showerror("Import Error", str(exc))

    def on_export_excel(self):  # "AI"
        """Export all products and transactions to an Excel file."""
        if not self.inventory.products:
            messagebox.showinfo("Export", "No products to export.")
            return
        default = f"warehouse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default, filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Warehouse Transactions"
            headers = ["Product", "Action", "Quantity", "Date/Time", "Stock After"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            for prod in self.inventory.list_products():
                current = 0
                for entry in prod.transactions:
                    act = entry["action"]
                    qty = entry["quantity"]
                    ts = entry["datetime"]
                    if act in ("initial", "initial_replace"):
                        current = qty
                    elif act == "add":
                        current += qty
                    elif act == "sell":
                        current -= qty
                    ws.append([prod.name, act, qty, ts, current])
            wb.save(path)
            messagebox.showinfo("Export", f"Excel saved:\n{path}")
        except Exception as exc:
            logger.exception("Failed to export excel: %s", exc)
            messagebox.showerror("Export Error", str(exc))

    # ---------------- Save / Load / Undo / Redo ----------------
    def _save_now(self):  # "AI"
        """Manually save inventory data to disk."""
        ok = self.inventory.save()
        if ok:
            messagebox.showinfo("Save", "Saved to disk.")
        else:
            messagebox.showerror("Save", "Failed to save. See console/log for details.")

    def on_undo(self):  # "AI"
        """Undo the last inventory action."""
        ok = self.inventory.undo()
        if not ok:
            messagebox.showinfo("Undo", "Nothing to undo.")
            return
        self.refresh_products_table()
        self.tree.selection_remove(self.tree.selection())
        messagebox.showinfo("Undo", "Undo applied.")

    def on_redo(self):  # "AI"
        """Redo the previously undone action."""
        ok = self.inventory.redo()
        if not ok:
            messagebox.showinfo("Redo", "Nothing to redo.")
            return
        self.refresh_products_table()
        messagebox.showinfo("Redo", "Redo applied.")

    def on_clear_all(self):  # "AI"
        """Clear all products, transactions, and undo/redo stacks."""
        if not messagebox.askyesno("Clear All", "Remove all products and transactions?"):
            return
        self.inventory.products.clear()
        self.inventory.undo_stack.clear()
        self.inventory.redo_stack.clear()
        self.inventory.save()
        self.refresh_products_table()
        messagebox.showinfo("Cleared", "All data removed.")

    def _manual_load(self):  # "AI"
        """Manually load saved inventory from disk, replacing in-memory data."""
        if not os.path.isfile(self.inventory.autosave_file):
            messagebox.showinfo("Load", "No saved data found.")
            return
        if not messagebox.askyesno("Load", "Load saved inventory from disk? This will replace current in-memory data."):
            return
        ok = self.inventory.load()
        if ok:
            self.refresh_products_table()
            messagebox.showinfo("Load", "Data loaded from disk.")
        else:
            messagebox.showerror("Load", "Failed to load. See console/log for details.")

def main():
    """Run WarehouseApp with a fresh Inventory instance."""
    inv = Inventory()
    app = WarehouseApp(inv)
    app.mainloop()


if __name__ == "__main__":
    main()

 #       #
 # #   # #
 #  # #  #
 #   #   #
 #       #

   #####   
  #     #  
 #       # 
 #       # 
  #     #  
   #####   

 #     #  
 #     #  
 #######  
 #     #  
 #     #  

    #     
   # #    
  #####   
 #     #  
######### 

 #       # 
 # #   # # 
 #  # #  # 
 #   #   # 
 #       # 

 #       # 
 # #   # # 
 #  # #  # 
 #   #   # 
 #       # 

    #     
   # #    
  #####   
 #     #  
######### 

######   
#     #  
#     #  
#     #  
######   


#####    
#    #   
#    #   
#####    
#   #    
#    #   

#####    
#        
#####    
#        
#####    

######    
    #    
   #     
  #      
######    

    #     
   # #    
  #####   
 #     #  
######### 